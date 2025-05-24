"""Class to create a report about the solution that was found"""

import pandas as pd
import pulp
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers, Alignment

from problemsolver import get_satisfaction_integral
import datareader


class SolutionAnalyzer:
    """Create a report about the solution found to the Linear Programming problem

    Which students were put together, how satisfied is everybody, which preferences
    were fulfilled, etc.
    """

    def __init__(
        self,
        fname,
        preferences: pd.DataFrame,
        input_sheet: pd.DataFrame,
        students_info: dict,
    ):
        self.fname = fname
        self.prob_vars, _ = pulp.LpProblem.from_json(fname)
        self.preferences = preferences
        self.input_sheet = input_sheet
        self.students_info = students_info

        self.groepsindeling = self._get_outcome()
        self.group_report = self._calculate_group_report()
        # The following calculations build upon eachother
        self.satisfied_constraints = self._calculate_satisfied_constraints()
        self.student_performance = self._calculate_performance_per_student()
        self.solution_performance = self._calculate_solution_performance()

    def _get_outcome(self) -> pd.DataFrame:
        """
        Restructure the Problem Variables in a nice DataFrame

        Parameters
        ----------
        vars: list of pulp.LpVariables
            The result of prob.variables()
        """
        chosen_groups = [
            var.name
            for name, var in self.prob_vars.items()
            if round(var.value()) == 1 and name.startswith("group")
        ]
        df = pd.DataFrame(chosen_groups)
        df[["Naam", "Group"]] = df[0].str.extract(r"group_\('(.*)',_'(.*)'\)")
        return df.drop(columns=[0])

    def display_transition_matrix(self):
        df_student_info = pd.DataFrame.from_dict(
            self.students_info, orient="index"
        ).reset_index(names="Naam")
        df = self.groepsindeling.merge(df_student_info)
        return pd.crosstab(df["Stamgroep"], df["Group"])

    def display_groepsindeling(self):
        """
        Transform DataFrame so that students are grouped by the group in which they are placed

        They are sorted by original Stamgroep. Below each new group, the number of jongens/meisjes
        and the total group size are shown
        """

        df_student_info = pd.DataFrame.from_dict(
            self.students_info, orient="index"
        ).reset_index(names="Naam")
        df = (
            self.groepsindeling.merge(df_student_info)
            .sort_values(["Jongen/meisje", "Stamgroep"])
            .assign(
                Naam=lambda df: df["Naam"] + " (" + df["Stamgroep"].str[:3] + ")",
                nr=lambda df: df.groupby(["Group", "Jongen/meisje"]).cumcount().add(1),
            )
            .set_index(["Group", "nr", "Jongen/meisje"])["Naam"]
            .unstack(["Group", "Jongen/meisje"], fill_value="")
            .sort_index(axis="columns")
        )

        # The double transpose works around a concat error for MultiIndex
        df = pd.concat(
            [df.transpose(), df.apply(lambda col: (col != "")).sum().rename("#")],
            axis="columns",
        ).transpose()

        for group in df.columns.levels[0]:
            df.loc["Groepsgrootte", (group, "Jongen")] = self.group_report.loc[
                (group, "Totaal"), "Groepsgrootte"
            ]

        return df

    def _calculate_group_report(self) -> pd.DataFrame:
        distribution = {}
        for geslacht in "Jongen", "Meisje":
            sex = "boys" if geslacht == "Jongen" else "girls"
            for gedeelte in "Totaal", "Jaarlaag":
                part = "in" if gedeelte == "Totaal" else "to"
                for group in self.groepsindeling["Group"].unique():
                    varname = f"{sex}_{part}_group_{group}"
                    distribution[(group, gedeelte, geslacht)] = round(
                        self.prob_vars[varname].value()
                    )

        df_group_report = (
            pd.Series(distribution)
            .unstack()
            .assign(
                VerschilJongensMeisjes=lambda df: (df["Jongen"] - df["Meisje"]).abs(),
                Groepsgrootte=lambda df: df["Jongen"] + df["Meisje"],
            )
            .astype(int)
        )

        return df_group_report

    @staticmethod
    def _probvars_to_series(prob_vars: dict, name: str, not_in_name: str) -> pd.Series:
        """
        Extract (accounted) preferences from problem to a series

        Will extract student name and preference number as index, and whether accounted
        for as value

        Parameters
        ----------
        prob_vars : dict
            Dictionary with name as key and LpVar as value
        name: str
            The beginning of the variable name, will also be the name of the Series
        not_in_name: str
            substring that can not appear in the variable name

        Returns
        -------
        pd.Series
            the values of the variables

        """
        constraints = {
            varname: var.value()
            for varname, var in prob_vars.items()
            if varname.startswith(name) and not not_in_name in varname
        }
        series = pd.Series(constraints, name=name)
        ix = (
            series.index.to_series()
            .str.extract(rf"{name}_\('(?P<student>.*)',_(?P<Nr>.*)\)")
            .set_index(["student", "Nr"])
            .index
        )

        series.index = ix
        return series

    def _calculate_satisfied_constraints(self) -> pd.DataFrame:
        """
        Calculate which constraints and for whom are accommodated

        Parameters
        ----------
        prob: pulp.LpProblem
            The result of prob.variables()

        Returns
        -------
            pd.DataFrame with Satisfied and WeightedSatisfied preferences
        """
        satisfied = (
            self._probvars_to_series(self.prob_vars, "Satisfied", "per_group")
            .astype(int)  # Floats must be converted to int before boolean accepts them
            .astype("boolean")
        )
        weighted_satisfied = self._probvars_to_series(
            self.prob_vars, "WeightedSatisfied", "per_group"
        )
        df = pd.concat([satisfied, weighted_satisfied], axis="columns")
        df.index = df.index.set_levels(pd.to_numeric(df.index.levels[1]), level=1)
        return df

    def _calculate_performance_per_student(self):
        """
        Calculate basic performance metrics per student

        Performance is better when more preferences are more accommodated

        Parameters
        ----------
        satisfied_constraints: pd.DataFrame
            The output of calculate_satisfied_constraints
        """
        df = (
            self.satisfied_constraints.groupby("student")
            .agg(
                NrPreferences=("Satisfied", "count"),
                AccountedPreferences=("Satisfied", "sum"),
                PctAccounted=("Satisfied", "mean"),
                AccountedWeightedPreferences=("WeightedSatisfied", "sum"),
            )
            .assign(
                NrWeightedPreferences=self.preferences.xs("Graag met", level="TypeWens")
                .loc[lambda df: df["Gewicht"].gt(0)]
                .groupby("Leerling")["Gewicht"]
                .sum(),
                PctWeightedAccounted=lambda df: df["AccountedWeightedPreferences"]
                / df["NrWeightedPreferences"],
                PossibleSatisfaction=lambda df: df["NrWeightedPreferences"].map(
                    lambda x: get_satisfaction_integral(0, x)
                ),
                ActualSatisfaction=lambda df: df["AccountedWeightedPreferences"].map(
                    lambda x: get_satisfaction_integral(0, x)
                ),
                RelativeSatisfaction=lambda df: df["ActualSatisfaction"]
                / df["PossibleSatisfaction"],
            )
        )
        return df

    def display_student_performance(self) -> pd.DataFrame:
        """Show the satisfaction per student as styled DataFrame

        Returns
        -------
        pd.DataFrame
            Table with information per student. Styled for optimal clarity
        """
        cols = {
            "RelativeSatisfaction": "Tevredenheid",
            "AccountedWeightedPreferences": "Aantal gehonoreerde wensen",
            "NrWeightedPreferences": "Aantal wensen",
        }
        return (
            self.student_performance.rename_axis("Leerling")
            .loc[:, list(cols.keys())]
            .rename(columns=cols)
            .style.background_gradient(
                "RdYlGn", vmin=0, vmax=1, subset=["Tevredenheid"]
            )
            .format(
                {
                    "Tevredenheid": "{:.2%}",
                    "Aantal gehonoreerde wensen": "{:.1f}",
                    "Aantal wensen": "{:.1f}",
                }
            )
        )

    def _calculate_solution_performance(self):
        """
        Calculate the performance of the general model
        """
        cols = [
            "NrPreferences",
            "AccountedPreferences",
            "NrWeightedPreferences",
            "AccountedWeightedPreferences",
            "PossibleSatisfaction",
            "ActualSatisfaction",
        ]
        solution_performance = (
            self.student_performance[cols]
            .sum()
            .to_frame()
            .transpose()
            .assign(
                PctAccountedPreferences=lambda df: df["AccountedPreferences"]
                / df["NrPreferences"],
                PctAccountedWeightedPreferences=lambda df: df[
                    "AccountedWeightedPreferences"
                ]
                / df["NrWeightedPreferences"],
                RelativeSatisfaction=lambda df: df["ActualSatisfaction"]
                / df["PossibleSatisfaction"],
            )
        ).to_dict("records")[0]
        return solution_performance

    def _determine_satisfied_preferences_studentindex(self) -> pd.DataFrame:
        """Get the satisfied preferences, but change the index so that it matches the input

        This is useful so that we can match the original file whether a preference is satisfied
        And is used in coloring the output

        Returns
        -------
        df
        """
        preferences_incl_liever_niet = datareader.toggle_negative_weights(
            self.preferences, mask="Gewicht"
        )
        mapping = {}
        for i in range(len(self.preferences)):
            if self.preferences.index[i][1] == "Graag met":
                mapping[self.preferences.reset_index("TypeWens").index[i]] = (
                    preferences_incl_liever_niet.index[i]
                )

        df = pd.DataFrame(mapping, index=["Leerling", "TypeWens", "Nr"]).transpose()
        df.index.names = ["student", "Nr"]

        df = (
            df.join(self.satisfied_constraints)
            .reset_index(drop=True)
            .set_index(["Leerling", "TypeWens", "Nr"])
        )
        return df

    @staticmethod
    def _display_satisfied_preferences(
        df: pd.DataFrame, satisfied_preferences_original_index: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Determine the background property based on whether a wish is satisfied

        Parameters
        ----------
        df : pd.DataFrame
            DataFrame that contains the right index and columns
        satisfied_preferences_original_index : pd.DataFrame
            DataFrame that contains wether the preference is satisfied. Index in long-form

        """
        df_style = pd.DataFrame(
            "background-color: white",
            index=df.index,
            columns=df.columns,
        )

        for idx in df_style.index:
            for col in df_style.columns:
                original_idx = (idx, col[0], col[1])
                try:
                    if satisfied_preferences_original_index.loc[
                        original_idx, "Satisfied"
                    ]:
                        df_style.loc[idx, col] = "background-color: green"
                    else:
                        df_style.loc[idx, col] = "background-color: red"
                except KeyError:  # Not a preference -> leave background as is
                    continue

        return df_style

    def display_satisfied_preferences(self) -> pd.DataFrame:
        """Display which preferences were satisfied and which werent in the original format

        Returns
        -------
        pd.DataFrame
            Style DataFrame for optimal clarity
        """
        satisfied_preferences_original_index = (
            self._determine_satisfied_preferences_studentindex()
        )
        return self.input_sheet.style.apply(
            self._display_satisfied_preferences,
            axis=None,
            satisfied_preferences_original_index=satisfied_preferences_original_index,
        )

    @staticmethod
    def _autoscale_column_width(sheet):
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = (
                adjusted_width
            )

    def to_excel(self, fname=None) -> None:
        """Put the most important outcomes of the solution in an Excel file

        Uses the three most important outcomes:
        - The acutal groepsindeling
        - The satisfaction per student
        - Which preferences were accounted for

        Each outcome is styled and shown in its own worksheet
        The solution metrics are not shown - they are probably too abstract for the
        end user

        Parameters
        ----------
        fname : str
            The filename to save to. By default the name of the json that was loaded
        """
        if fname is None:
            fname = self.fname.replace("json", "xlsx")
        # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
        with pd.ExcelWriter(fname, engine="openpyxl") as writer:
            self._write_groepsindeling(writer)
            self.group_report.to_excel(writer, sheet_name="Klassenoverzicht")
            self.display_transition_matrix().to_excel(
                writer, sheet_name="Overgangsmatrix"
            )

            self.display_student_performance().to_excel(
                writer, sheet_name="Leerlingtevredenheid"
            )
            sheet = writer.book.worksheets[1]
            for cell in sheet["B"]:
                cell.number_format = numbers.FORMAT_PERCENTAGE
            self._autoscale_column_width(sheet)

            self.display_satisfied_preferences().to_excel(
                writer, sheet_name="VervuldeWensen"
            )

    def _write_groepsindeling(self, writer):
        groepsindeling = self.display_groepsindeling()
        groepsindeling.iloc[:-1].to_excel(writer, sheet_name="Groepsindeling")
        sheet = writer.sheets["Groepsindeling"]

        row = (
            len(groepsindeling) + len(groepsindeling.columns.levels) + 1
        )  # Excel is 1-based + header
        col_index = 2  # Start bij kolom B in Excel (A is index)
        for group in groepsindeling.columns.levels[0]:
            sheet.merge_cells(
                start_row=row,
                start_column=col_index,
                end_row=row,
                end_column=col_index + 1,
            )
            sheet.cell(row=row, column=col_index).value = groepsindeling.loc[
                "Groepsgrootte", (group, "Jongen")
            ]
            sheet.cell(row=row, column=col_index).alignment = Alignment("center")
            col_index += 2
        self._autoscale_column_width(sheet)

    def get_hash(self) -> int:
        "Give unique code for groepsindeling"
        return hash(tuple(self._get_outcome().sort_values("Naam")["Group"]))

"""The flask server that governs the app"""

import logging
import os
import uuid
import webbrowser
import zipfile
from collections import Counter, defaultdict
from io import BytesIO
from threading import Thread

import numpy as np
import openpyxl
import pandas as pd
import pandera as pa
from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl.worksheet.datavalidation import DataValidation

from aliexpress import datareader, sociogram
from aliexpress.errors import (
    CouldNotReadFileError,
    FeasibilityError,
    ValidationError,
)
from aliexpress.main import distribute_students_once


def setup_logger():
    """Create logging instance"""
    log = logging.getLogger(__name__)
    log.setLevel(logging.INFO)

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler("aliexpress.log")
    file_handler.setLevel(logging.DEBUG)

    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    log.addHandler(file_handler)
    log.addHandler(console_handler)
    return log


logger = setup_logger()


load_dotenv()

env = os.getenv("FLASK_ENV", "production")
if env == "development":
    from src.aliexpress.appconfig import DevelopmentConfig as ConfigClass
else:
    from src.aliexpress.appconfig import ProductionConfig as ConfigClass

app = Flask(__name__)
app.config.from_object(ConfigClass)


temp_storage = {}
status_dct = defaultdict(
    lambda: {
        "status_studentdistribution": "pending",
        "status_sociogram": "pending",
        "logs": [],
    }
)


@app.route("/")
def home():
    """Display home page"""
    return render_template("home.html")


def file_to_io(uploaded_file) -> BytesIO:
    """Get file as BytesIO"""
    return BytesIO(uploaded_file.read())


@app.route("/input_templates/<path:filename>")
def download_template(filename):
    """Download the template sheets"""
    return send_from_directory("input_templates", filename, as_attachment=True)


@app.route("/fillin", methods=["GET", "POST"])
def fillin():
    """Display and process the fillin page"""
    if request.method == "POST":
        if "edexml" in request.files:
            return _handle_edexml_upload(request)
        else:
            return _handle_form_submission(request)

    return render_template("fillin.html")


def _handle_edexml_upload(req):
    """Process uploaded edexml and render candidates + groups"""
    edexml = file_to_io(req.files["edexml"])
    jaargroep = int(req.form["jaargroep"])
    df = datareader.EdexReader(edexml).get_full_df()

    candidates = (
        df.loc[lambda df: df["jaargroep"] == jaargroep]
        .sort_values(["groepsnaam", "roepnaam", "achternaam"])
        .reset_index()
        .filter(
            ["key", "roepnaam", "achternaam", "groepsnaam", "geslacht"], axis="columns"
        )
        .to_dict(orient="records")
    )
    # Later the values will be retrieved based on key (which are selected in the app)
    temp_storage["candidates"] = candidates

    groups_from = (
        df.loc[lambda df: df["jaargroep"] == jaargroep, "groepsnaam"].unique().tolist()
    ) + ["Anders"]

    groupnames_to = (
        df.loc[lambda df: df["jaargroep"] == jaargroep + 1, "groepsnaam"]
        .unique()
        .tolist()
    )

    groups_to = (
        df.loc[lambda df: df["groepsnaam"].isin(groupnames_to)]
        .assign(
            blijft_in_groep=lambda df: df["jaargroep"]
            < df.groupby("groepsnaam")["jaargroep"].transform("max")
        )
        .sort_values(["groepsnaam", "jaargroep", "geslacht"])
        .groupby("groepsnaam")
        .apply(
            lambda g: g[
                ["roepnaam", "achternaam", "geslacht", "jaargroep", "blijft_in_groep"]
            ].to_dict(orient="records")
        )
        .to_dict()
    )

    return render_template(
        "fillin.html",
        candidates=candidates,
        groups_from=groups_from,
        groups_to=groups_to,
        uploaded=True,
    )


def _handle_form_submission(req):
    """Process the form after candidates have been selected and groups defined"""
    formdata = req.form.to_dict(flat=False)

    selected_per_group = _extract_selected_per_group(formdata)
    groups_to = _build_groups_summary(selected_per_group, formdata)

    selected_ids = req.form.getlist("students")
    new_students = _extract_new_students(req)

    df_total = _combine_students(temp_storage["candidates"], selected_ids, new_students)

    zip_buffer = _create_zip_with_templates(groups_to, df_total)

    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name="invulformulieren.zip",
        mimetype="application/zip",
    )


def _extract_selected_per_group(formdata):
    """Extract students assigned to each group from formdata"""
    selected = defaultdict(list)
    for key, values in formdata.items():
        if key.startswith("group_students["):
            groupname = key[len("group_students[") : -1]  # extract text inside [ ]
            selected[groupname].extend(values)
    return selected


def _build_groups_summary(selected_per_group, formdata):
    """Build a summary of groups with counts of boys and girls"""
    groups_to = {}
    for g, lst in selected_per_group.items():
        c = Counter(lst)
        groups_to[g] = {"Jongens": c.get("Jongen", 0), "Meisjes": c.get("Meisje", 0)}

    for g in [grp for grp in formdata.get("new_groups[]", []) if grp.strip()]:
        groups_to.setdefault(g, {"Jongens": 0, "Meisjes": 0})

    return groups_to


def _extract_new_students(req):
    """Extract manually added students from form fields"""
    firstnames = req.form.getlist("new_firstname[]")
    lastnames = req.form.getlist("new_lastname[]")
    genders = req.form.getlist("new_gender[]")
    groups = req.form.getlist("new_group[]")

    return [
        {"roepnaam": fn, "achternaam": ln, "geslacht": sex, "groepsnaam": gr}
        for fn, ln, sex, gr in zip(firstnames, lastnames, genders, groups)
        if fn.strip() and ln.strip()
    ]


def _combine_students(candidates, selected_ids, new_students):
    """Combine selected and new students into a single DataFrame"""
    df_original = pd.DataFrame(candidates).set_index("key").loc[selected_ids]
    df_new = pd.DataFrame(new_students)
    return (
        pd.concat([df_original, df_new])
        .assign(uniekenaam=create_unique_name)
        .sort_values(["groepsnaam", "uniekenaam"])
    )


def _create_zip_with_templates(groups_to, df_total):
    """Fill in Excel templates and package them into a ZIP file"""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

        wb_groups = openpyxl.load_workbook("input_templates/groepen.xlsx")
        fill_in_groups_to(groups_to, wb_groups)
        _add_to_zip(zip_file, wb_groups, "groepen.xlsx")

        wb_prefs = openpyxl.load_workbook("input_templates/voorkeuren.xlsx")
        fill_in_known_values(list(groups_to.keys()), df_total, wb_prefs)
        add_data_validations(wb_prefs)
        _add_to_zip(zip_file, wb_prefs, "voorkeuren.xlsx")

        wb_not_together = openpyxl.load_workbook("input_templates/niet_samen.xlsx")
        add_data_validations_not_together(wb_not_together, df_total)
        _add_to_zip(zip_file, wb_not_together, "niet_samen.xlsx")

    zip_buffer.seek(0)
    return zip_buffer


def _add_to_zip(zip_file, workbook, filename):
    """Helper to save workbook to buffer and add to ZIP"""
    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    zip_file.writestr(filename, buf.read())


def create_unique_name(df: pd.DataFrame) -> pd.Series:
    """Find unique name per leerling. Needs roepnaam and achternaam"""
    unique_names = df["roepnaam"] + " "

    n_letters_added = 0
    while unique_names.duplicated().any():
        for ix in unique_names[unique_names.duplicated(keep=False)].index:
            unique_names[ix] += df.loc[ix, "achternaam"][n_letters_added]
        n_letters_added += 1
    return unique_names.str.strip()


def add_data_validations_not_together(wb, df):
    """Add data validations for students to not_together"""
    ws2 = wb["Sheet2"]
    for i, ll in enumerate(df["uniekenaam"].unique().tolist(), start=1):
        ws2[f"A{i}"].value = ll

    ws1 = wb["Sheet1"]

    dv = DataValidation(
        type="list",
        formula1="=Sheet2!$A:$A",
        allow_blank=True,
        showErrorMessage=True,
    )
    for col in "BCDEFGHIJKLM":
        dv.add(f"{col}2:{col}1048576")
    ws1.add_data_validation(dv)


def add_data_validations(wb):
    """Add data validations to workbook

    For jongen/meisje, for niet in (groups) and for preferences (students + group)
    """
    ws1 = wb["Sheet1"]
    val_specs = [
        ("Sheet2!$A:$A", "C"),  # jongen/meisje
        ("Sheet2!$B:$B", "QR"),  # niet in: only groups
        ("Sheet2!$C:$C", "EGIKMO"),  # (negative) preferences: groups + students
    ]

    for rng, cols_to_be_validated in val_specs:
        dv = DataValidation(
            type="list", formula1=f"={rng}", allow_blank=True, showErrorMessage=True
        )
        for col in cols_to_be_validated:
            dv.add(f"{col}4:{col}1048576")
        ws1.add_data_validation(dv)


def fill_in_groups_to(groups_to, wb):
    """Fill the students in from the workbook, and the data to be used for validation"""
    ws1 = wb["Sheet1"]
    for i, (gr, values) in enumerate(groups_to.items(), start=2):
        ws1[f"A{i}"] = gr
        ws1[f"B{i}"] = values["Jongens"]
        ws1[f"C{i}"] = values["Meisjes"]

        ws1[f"A{i}"].protection = openpyxl.styles.Protection(locked=False)
        ws1[f"B{i}"].protection = openpyxl.styles.Protection(locked=False)
        ws1[f"C{i}"].protection = openpyxl.styles.Protection(locked=False)

    dv_int = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Alleen niet-negatieve gehele getallen zijn toegestaan.",
    )
    dv_int.add("B2:B1048576")
    dv_int.add("C2:C1048576")
    ws1.add_data_validation(dv_int)
    ws1.protection.sheet = True


def fill_in_known_values(groups_to, groep_die_doorgaat, wb):
    """Fill the students in from the workbook, and the data to be used for validation"""
    ws1 = wb["Sheet1"]
    for i, (_, row) in enumerate(groep_die_doorgaat.iterrows(), start=4):
        ws1[f"A{i}"].value = row["uniekenaam"]
        ws1[f"C{i}"].value = row["geslacht"]
        ws1[f"D{i}"].value = row["groepsnaam"]

    logger.debug("Data ingevuld")

    all_leerlingen = groep_die_doorgaat["uniekenaam"].tolist()
    ws2 = wb["Sheet2"]
    for i, gr in enumerate(groups_to, start=1):
        ws2[f"B{i}"].value = gr
    for i, sub in enumerate(groups_to + all_leerlingen, start=1):
        ws2[f"C{i}"].value = sub


def to_validation_message(exc: Exception) -> str:
    """Convert a validation exception to a user-friendly message"""
    if isinstance(exc, pa.errors.SchemaError):
        return schemaerror_to_validation_message(exc)
    if isinstance(exc, (ValidationError, CouldNotReadFileError, FeasibilityError)):
        return readableerror_to_validation_message(exc)
    return (
        "Er is iets onverwachts misgegaan. Het probleem is gelogd. "
        "Laat de maker dit onderzoeken."
    )


def readableerror_to_validation_message(exc: Exception) -> str:
    """Convert a validation exception to a user-friendly message"""
    friendly_templates = {
        "wrong_columns_preferences": (
            "Het voorkeuren-bestand heeft de verkeerde kolommen. Controleer of je het goede"
            " bestand hebt geupload en het meest recente template hebt gebruikt. "
            "\n{wrong_columns}"
        ),
        "infeasible_problem": (
            "Met deze vereiste klassenbalans en verdeling van leerlingen die overgaan is het"
            "niet mogelijk. Overweeg de volgende versoepelingen om het probleem wel op te "
            "lossen:\n {possible_improvement}"
        ),
        "internal_error": (
            "Er is iets onverwachts misgegaan. Het probleem is gelogd. "
            "Laat de maker dit onderzoeken."
        ),
    }

    template = friendly_templates.get(exc.code, None)
    if template:
        return template.format(**exc.context)
    return (
        "Er is iets onverwachts misgegaan. Het probleem is gelogd. "
        "Laat de maker dit onderzoeken."
    )


def schemaerror_to_validation_message(exc: pa.errors.SchemaError) -> str:
    """Convert a pandera SchemaError to a user-friendly message

    This SchemaError must have been modified to contain a 'filetype' attribute.
    """
    if exc.reason_code in (
        pa.errors.SchemaErrorReason.COLUMN_NOT_IN_SCHEMA,
        pa.errors.SchemaErrorReason.COLUMN_NOT_IN_DATAFRAME,
    ):
        return (
            f"Het {exc.filetype}-bestand heeft de verkeerde kolommen. Controleer of je het goede"
            " bestand hebt geupload en het meest recente template hebt gebruikt. "
            f"\n{exc.failure_cases}"
        )
    if exc.reason_code == pa.errors.SchemaErrorReason.DATATYPE_COERCION:
        return (
            f"Ongeldige waarden gevonden in kolom {exc.schema.name} "
            f"van het {exc.filetype}-bestand"
        )
    if exc.reason_code == pa.errors.SchemaErrorReason.SERIES_CONTAINS_NULLS:
        return (
            f"In het {exc.filetype}-bestand zijn niet alle verplichte kolommen gevuld: "
            f"controleer {exc.column_name} bij regel "
            f"{', '.join(exc.failure_cases[', '].astype(str))}"
        )
    if exc.reason_code == pa.errors.SchemaErrorReason.SERIES_CONTAINS_DUPLICATES:
        if exc.filetype == "voorkeuren":
            duplicates = ", ".join(exc.failure_cases["failure_case"])
            return (
                f"In voorkeuren is de volgende naam/namen niet uniek: {duplicates}\n"
                "Voeg de eerste letter van de achternaam toe om de leerlingen van "
                "elkaar te onderscheiden."
            )
        return (
            f"In het {exc.filetype}-bestand zijn dubbelingen ingevuld "
            f"in kolom {exc.column_name}"
        )

    if exc.reason_code == pa.errors.SchemaErrorReason.DATAFRAME_CHECK:
        if exc.check.name == "empty_df":
            return (
                f"Het {exc.filetype}-bestand was helemaal leeg. Daardoor kan er "
                "geen groepsindeling worden berekend"
            )
        if exc.column_name == ("Jongen/meisje", np.nan, np.nan):
            return f"Verkeerd ingevuld geslacht voor {', '.join(exc.failure_cases['index'])}"
        if exc.check.name == "greater_than" and "Gewicht" in exc.column_name:
            return "Er zijn negatieve gewichten in het voorkeurenbestand."
        if exc.check.name == "duplicated_values_preferences":
            students_with_duplicates = ", ".join(
                set(exc.failure_cases["index"].get_level_values("Leerling"))
            )
            return (
                "In het voorkeuren-bestand is voor "
                f"{students_with_duplicates} een leerling of groep gevonden die "
                "dubbel voorkomt. Tel ze op of streep ze tegen elkaar weg om "
                "dubbelingen te voorkomen."
            )
        if exc.check.name == "invalid_values_preferences":
            invalid_values = ", ".join(
                set(
                    exc.failure_cases.loc[
                        lambda df: df["column"] == "Waarde", "failure_case"
                    ]
                )
            )
            return f"Onbekende leerling of groep in categorie: {invalid_values}"
        if exc.check.name == "isin" and exc.filetype == "niet_samen":
            unknown_students = ", ".join(exc.failure_cases["failure_case"].astype(str))
            return (
                f"In het niet-samen-bestand komt {unknown_students} voor, "
                "die niet in het voorkeurenbestand voorkomt"
            )
        if exc.check.name == "duplicated_students_not_together":
            rows = ", ".join(set(exc.failure_cases["index"].add(1).astype(str)))
            duplicated_students = ", ".join(
                exc.failure_cases.groupby("index")["failure_case"].apply(
                    lambda s: s[s.duplicated()]
                )
            )
            return (
                f"In het niet-samen-bestand wordt in de {rows}e "
                f"groep dezelfde leerling meerdere keren genoemd: {duplicated_students}"
            )
        if exc.check.name == "too_strict_not_together":
            rows = ", ".join(set(exc.failure_cases["index"].add(1).astype(str)))
            max_samen = ", ".join(
                exc.failure_cases.loc[
                    lambda df: df["column"] == "Max aantal samen", "failure_case"
                ].astype(str)
            )
            nr_students = ", ".join(
                exc.failure_cases.groupby("index").size().sub(1).astype(str)
            )

            return (
                f"In het niet-samen-bestand op de {rows}e rij is de maximale "
                f"groepsgrootte te klein: met dit aantal groepen lukt het niet om {nr_students} "
                f"leerlingen te verdelen met maximaal {max_samen} bij elkaar."
            )

    return (
        f"Er is iets onverwachts misgegaan bij het lezen van {exc.filetype}. "
        "Controleer het bestand goed en of je het meest recente template hebt gebruikt. "
        "Als het probleem blijft bestaan, laat de maker dit onderzoeken."
    )


def _handle_failure(exc, task_id, log_msg):
    logger.exception(log_msg)
    message = to_validation_message(exc)
    status_dct[task_id]["status_studentdistribution"] = "error"
    status_dct[task_id]["message"] = message


@app.route("/upload", methods=["GET", "POST"])
def upload_files():
    """Handle upload page, including form submission"""
    if request.method == "POST":
        logger.info("Submitted")
        preferences = file_to_io(request.files["preferences"])
        groups_to = file_to_io(request.files["groups_to"])
        not_together = file_to_io(request.files["not_together"])

        try:
            max_diff_n_students_total = int(request.form["max_diff_n_students_total"])
            max_diff_n_students_year = int(request.form["max_diff_n_students_year"])
            max_imbalance_boys_girls_total = int(
                request.form["max_imbalance_boys_girls_total"]
            )
            max_imbalance_boys_girls_year = int(
                request.form["max_imbalance_boys_girls_year"]
            )
            max_clique = int(request.form["max_clique"])
            max_clique_sex = int(request.form["max_clique_sex"])
        except (KeyError, ValueError):
            return "Alle parameters moeten positieve gehele getallen zijn", 400

        kwargs = {
            "max_diff_n_students_total": max_diff_n_students_total,
            "max_diff_n_students_year": max_diff_n_students_year,
            "max_imbalance_boys_girls_total": max_imbalance_boys_girls_total,
            "max_imbalance_boys_girls_year": max_imbalance_boys_girls_year,
            "max_clique": max_clique,
            "max_clique_sex": max_clique_sex,
        }
        session["config"] = kwargs

        def on_update(message):
            status_dct[task_id]["logs"].append(message)

        logger.info("Starting distribution...")

        task_id = str(uuid.uuid4())
        temp_storage[task_id] = {}

        # pylint: disable=broad-exception-caught
        def run_task(*args, **kwargs):
            try:
                status_dct[task_id]["status_studentdistribution"] = "running"
                result = distribute_students_once(*args, **kwargs, on_update=on_update)
                logger.info("Distributing students finished successfully")
                status_dct[task_id]["status_studentdistribution"] = "done"
                temp_storage[task_id]["groepsindeling"] = result
            except (
                pa.errors.SchemaError,
                ValidationError,
                CouldNotReadFileError,
            ) as exc:
                _handle_failure(exc, task_id, "Files are incorrect")
            except FeasibilityError as exc:
                _handle_failure(exc, task_id, "Problem is infeasible")
            except Exception as exc:
                _handle_failure(exc, task_id, "Uncaught exception")

        def create_sociogram(preferences, groups_to):
            try:
                on_update("Sociogram tekenen...")
                groups_to = list(datareader.read_groups_excel(groups_to).keys())
                sg = sociogram.SociogramMaker(preferences, groups_to)
                fig, g, pos = sg.plot_sociogram()
                logger.info("Sociogram created")

                fig = sociogram.networkx_to_plotly(g, pos)
                html = fig.to_html(full_html=False, include_plotlyjs="cdn")
                logger.info("HTML created")
                on_update(
                    f'<a href=/sociogram/{task_id} target="_blank" class="button">'
                    "Bekijk het sociogram nu!</a>"
                )
                temp_storage[task_id]["sociogram"] = html
            except Exception:
                logger.exception("Could not create sociogram")

        # pylint: enable=broad-exception-caught
        Thread(target=create_sociogram, args=(preferences, groups_to)).start()
        Thread(
            target=run_task,
            args=(preferences, groups_to, not_together),
            kwargs=kwargs,
        ).start()

        return redirect(url_for("processing", task_id=task_id))
    logger.info("Showing upload page")
    return render_template("upload.html")


@app.route("/status/<task_id>")
def status(task_id):
    """Return status as json"""
    result = status_dct.get(task_id)
    if not result:
        return jsonify({"status_studentdistribution": "unknown"})
    return jsonify(result)


@app.route("/processing/<task_id>")
def processing(task_id):
    """Display processing page"""
    return render_template("processing.html", task_id=task_id)


@app.route("/handle-error", methods=["POST"])
def handle_error():
    """Show information about errors to user"""
    data = request.get_json()
    flash(data["message"], "error")

    # By not redirecting here but in JS, this is more flexible
    return "", 204


@app.route("/sociogram/<task_id>")
def show_sociogram(task_id):
    """Display sociogram"""
    html = temp_storage[task_id]["sociogram"]
    return render_template("sociogram.html", plotly_div=html)


@app.route("/result/<task_id>")
def result_page(task_id):
    """Display result for single run"""

    dataframes = {
        k: df.to_html(na_rep="")
        for k, df in temp_storage[task_id]["groepsindeling"]["dataframes"].items()
    }
    return render_template("result.html", task_id=task_id, dataframes=dataframes)


@app.route("/download/<task_id>")
def download(task_id):
    """Download single groepsindeling"""
    file_buffer = temp_storage.get(task_id)
    logger.debug(task_id)
    if file_buffer is None:
        flash("Groepsindeling niet gevonden. Mogelijk nog aan het berekenen", "error")
        return render_template("result.html", task_id=task_id)

    return send_file(
        file_buffer["groepsindeling"]["download"],
        as_attachment=True,
        download_name="results.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/done")
def done():
    """Show done page"""
    return render_template("done.html")


if __name__ == "__main__":
    webbrowser.open("http://localhost:5000")
    app.run(debug=True)

"""Class to create a report about the solution that was found"""

import dataclasses

import pandas as pd
from IPython.display import display
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter

from . import datareader

TABLE_STYLES = styles = [
    {
        "selector": "th.row_heading, td.row_heading, th.row_heading.level0",
        "props": [
            (
                "background-color",
                "#f0f0f0",
            ),  # light grey background for index cells
            ("border", "1px solid #dcdcdc"),
            ("padding", "6px 10px"),
        ],
    },
    {
        "selector": "table.dataframe",
        "props": [
            ("border-collapse", "collapse"),
            ("border", "1px solid #dcdcdc"),
        ],
    },
    {
        "selector": "th.col_heading",
        "props": [
            (
                "background-color",
                "#e9ecef",
            ),  # light grey for column headers (optional)
            ("border", "1px solid #dcdcdc"),
            ("padding", "6px 10px"),
        ],
    },
]


@dataclasses.dataclass(frozen=True)
class DisplayNames:
    """Maps matching keys back to the names as entered, per namespace, for reporting.

    The solver works on matching keys; these maps let the report layer show students,
    target groups and current groups (Stamgroep) exactly as the user typed them. Each
    defaults to empty, in which case keys are shown unchanged.
    """

    student: dict = dataclasses.field(default_factory=dict)
    group: dict = dataclasses.field(default_factory=dict)
    stamgroep: dict = dataclasses.field(default_factory=dict)


# pylint: disable-next=too-many-instance-attributes  # ten computed views of one solution; each is a distinct output table
class SolutionAnalyzer:
    """Create a report about the solution found to the Linear Programming problem

    Which students were put together, how satisfied is everybody, which preferences
    were fulfilled, etc.
    """

    # Takes the solver result plus the three original-data views and an optional
    # display-name bundle; revisit for a cleaner split (see avoid-pylint-disable note).
    # pylint: disable-next=too-many-arguments,too-many-positional-arguments
    def __init__(
        self,
        result,
        preferences: pd.DataFrame,
        input_sheet: pd.DataFrame,
        students_info: dict,
        display_names: "DisplayNames | None" = None,
    ):
        # The solver works on matching keys; everything below this point is in display
        # space, so we relabel the solver output, the input sheet, the preferences and the
        # student metadata up front and leave every view method to work on names as the
        # user typed them.
        display_names = display_names or DisplayNames()
        self.student_display = display_names.student
        self.group_display = display_names.group
        self.stamgroep_display = display_names.stamgroep

        self.result = self._relabel_result(result)
        self.preferences = self._relabel_preferences(preferences)
        self.input_sheet = self._relabel_input_sheet(input_sheet)
        self.students_info = self._relabel_students_info(students_info)

        self.groepsindeling = self._get_outcome()
        self.group_report = self._calculate_group_report()
        # The following calculations build upon eachother
        self.satisfied_constraints = self._calculate_satisfied_constraints()
        self.student_performance = self._calculate_performance_per_student()
        self.solution_performance = self._calculate_solution_performance()

    def _student_name(self, key):
        """Display name for a student matching key."""
        return self.student_display.get(key, key)

    def _group_name(self, key):
        """Display name for a group matching key."""
        return self.group_display.get(key, key)

    def _relabel_result(self, result):
        """Rewrite the solver result from matching keys to names as entered."""
        return dataclasses.replace(
            result,
            assignment={
                self._student_name(s): self._group_name(g)
                for s, g in result.assignment.items()
            },
            student_satisfaction={
                self._student_name(s): v for s, v in result.student_satisfaction.items()
            },
            satisfied={
                (self._student_name(s), nr): v
                for (s, nr), v in result.satisfied.items()
            },
            weighted_satisfied={
                (self._student_name(s), nr): v
                for (s, nr), v in result.weighted_satisfied.items()
            },
            weights={
                (self._student_name(s), nr): v for (s, nr), v in result.weights.items()
            },
            group_composition={
                self._group_name(g): comp
                for g, comp in result.group_composition.items()
            },
        )

    def _relabel_preferences(self, preferences: pd.DataFrame) -> pd.DataFrame:
        """Relabel the Leerling index level to match the display-keyed result."""
        if "Leerling" not in preferences.index.names:
            return preferences
        new = preferences.copy()
        new.index = pd.MultiIndex.from_arrays(
            [
                new.index.get_level_values("Leerling").map(self._student_name),
                new.index.get_level_values("TypeWens"),
                new.index.get_level_values("Nr"),
            ],
            names=["Leerling", "TypeWens", "Nr"],
        )
        return new

    def _relabel_students_info(self, students_info: dict) -> dict:
        """Relabel student keys and their Stamgroep value to names as entered."""
        relabeled = {}
        for student, info in students_info.items():
            info = dict(info)
            if "Stamgroep" in info:
                info["Stamgroep"] = self.stamgroep_display.get(
                    info["Stamgroep"], info["Stamgroep"]
                )
            relabeled[self._student_name(student)] = info
        return relabeled

    def _relabel_input_sheet(self, input_sheet: pd.DataFrame) -> pd.DataFrame:
        """Relabel the original input sheet (index + name cells) to names as entered.

        Column-aware: 'Niet in' targets are groups, 'Graag met'/'Liever niet met' targets
        are a classmate or a group, and the Stamgroep column is a current group.
        """
        df = input_sheet.copy()
        df.index = df.index.map(self._student_name)
        student_or_group = {**self.group_display, **self.student_display}
        for col in df.columns:
            type_wens = col[0] if isinstance(col, tuple) else col
            type_waarde = col[2] if isinstance(col, tuple) and len(col) > 2 else None
            if type_wens == "Stamgroep":
                df[col] = df[col].map(lambda v: self.stamgroep_display.get(v, v))
            elif type_waarde == "Waarde" and type_wens == "Niet in":
                df[col] = df[col].map(lambda v: self.group_display.get(v, v))
            elif type_waarde == "Waarde" and type_wens in (
                "Graag met",
                "Liever niet met",
            ):
                df[col] = df[col].map(lambda v: student_or_group.get(v, v))
        return df

    def _get_outcome(self) -> pd.DataFrame:
        """Restructure the student -> group assignment into a [Naam, Group] DataFrame."""
        return (
            pd.Series(self.result.assignment, name="Group")
            .rename_axis("Naam")
            .reset_index()
        )

    @staticmethod
    def _indexed_series(mapping: dict, name: str) -> pd.Series:
        """Build a (student, Nr)-indexed Series from a {(student, Nr): value} mapping.

        Handles the empty case (no positive preferences) without losing the index names
        that the per-student aggregations rely on.
        """
        series = pd.Series(mapping, name=name)
        if len(series):
            series.index = series.index.set_names(["student", "Nr"])
        else:
            series.index = pd.MultiIndex.from_arrays([[], []], names=["student", "Nr"])
        return series

    def display_transition_matrix(self):
        """Create a transition matrix of the groups

        How many students moved from one group to another"""
        df_student_info = pd.DataFrame.from_dict(
            self.students_info, orient="index"
        ).reset_index(names="Naam")
        df = self.groepsindeling.merge(df_student_info)
        return pd.crosstab(df["Stamgroep"], df["Group"])

    def display_groepsindeling(self):
        """
        Transform DataFrame so that students are grouped by the group in which they are placed

        They are sorted by original Stamgroep. Below each new group, the number of jongens/meisjes
        and the total group size are shown
        """

        df_student_info = pd.DataFrame.from_dict(
            self.students_info, orient="index"
        ).reset_index(names="Naam")

        df = (
            self.groepsindeling.merge(df_student_info)
            .sort_values(["Jongen/meisje", "Stamgroep"])
            .assign(
                Naam=lambda df: df["Naam"] + " (" + df["Stamgroep"].str[:3] + ")",
                nr=lambda df: df.groupby(["Group", "Jongen/meisje"]).cumcount().add(1),
            )
            .set_index(["Group", "nr", "Jongen/meisje"])["Naam"]
            .unstack(["Group", "Jongen/meisje"], fill_value="")
        )

        # Show all columns even if some groups/sexes are not distributed to
        # pylint: disable=unsubscriptable-object
        expected_columns = pd.MultiIndex.from_product(
            [sorted(self.groepsindeling["Group"].unique()), ["Jongen", "Meisje"]],
            names=["Groep", "Jongen/meisje"],
        )

        # The double transpose works around a concat error for MultiIndex
        df = (
            pd.concat(
                [df.transpose(), df.apply(lambda col: (col != "")).sum().rename("#")],
                axis="columns",
            )
            .transpose()
            .reindex(expected_columns, axis="columns")
        )

        for group in df.columns.levels[0]:
            df.loc["Groepsgrootte", (group, "Jongen")] = self.group_report.loc[
                (group, "Totaal"), "Groepsgrootte"
            ]

        return df

    def _calculate_group_report(self) -> pd.DataFrame:
        distribution = {}
        for group, comp in self.result.group_composition.items():
            distribution[(group, "Totaal", "Jongen")] = comp.boys_total
            distribution[(group, "Totaal", "Meisje")] = comp.girls_total
            distribution[(group, "Jaarlaag", "Jongen")] = comp.boys_year
            distribution[(group, "Jaarlaag", "Meisje")] = comp.girls_year

        df_group_report = (
            pd.Series(distribution)
            .unstack()
            .assign(
                VerschilJongensMeisjes=lambda df: (df["Jongen"] - df["Meisje"]).abs(),
                Groepsgrootte=lambda df: df["Jongen"] + df["Meisje"],
            )
            .astype(int)
        )

        return df_group_report

    def _calculate_satisfied_constraints(self) -> pd.DataFrame:
        """Per (student, Nr): whether the wish is satisfied and its weighted value.

        Returns
        -------
            pd.DataFrame with Satisfied (boolean) and WeightedSatisfied preferences
        """
        satisfied = self._indexed_series(self.result.satisfied, "Satisfied").astype(
            "boolean"
        )
        weighted_satisfied = self._indexed_series(
            self.result.weighted_satisfied, "WeightedSatisfied"
        )
        return pd.concat([satisfied, weighted_satisfied], axis="columns")

    def _calculate_performance_per_student(self):
        """
        Calculate basic performance metrics per student

        Performance is better when more preferences are more accommodated
        """
        studentsatisfaction = pd.Series(
            self.result.student_satisfaction, name="RelativeSatisfaction"
        )
        n_weighted_preferencs = (
            self._indexed_series(self.result.weights, "Weights_preferences")
            .where(lambda s: s.gt(0))
            .groupby("student")
            .sum()
            .rename("NrWeightedPreferences")
        )
        df = (
            self.satisfied_constraints.groupby("student")
            .agg(
                NrPreferences=("Satisfied", "count"),
                AccountedPreferences=("Satisfied", "sum"),
                PctAccounted=("Satisfied", "mean"),
                AccountedWeightedPreferences=("WeightedSatisfied", "sum"),
            )
            .join(studentsatisfaction, how="outer")
            .join(n_weighted_preferencs)
            .fillna(0)
        )
        return df

    def display_student_performance(self) -> pd.DataFrame:
        """Show the satisfaction per student as styled DataFrame

        Returns
        -------
        pd.DataFrame
            Table with information per student. Styled for optimal clarity
        """
        cols = {
            "RelativeSatisfaction": "Tevredenheid",
            "AccountedWeightedPreferences": "Aantal gehonoreerde wensen",
            "NrWeightedPreferences": "Aantal wensen",
        }

        # styled = df.style

        return (
            self.student_performance.rename_axis("Leerling")
            .loc[:, list(cols.keys())]
            .rename(columns=cols)
            .style.background_gradient(
                "RdYlGn", vmin=0, vmax=1, subset=["Tevredenheid"]
            )
            .format(
                {
                    "Tevredenheid": "{:.2%}",
                    "Aantal gehonoreerde wensen": "{:.1f}",
                    "Aantal wensen": "{:.1f}",
                },
                na_rep="",
            )
            # Student names live in the index and are rendered raw via | safe; escape them.
            .format_index(escape="html", axis="index")
            .set_table_styles(TABLE_STYLES)
        )

    def _calculate_solution_performance(self):
        """
        Calculate the performance of the general model
        """
        cols = [
            "NrPreferences",
            "NrWeightedPreferences",
            "AccountedPreferences",
            "AccountedWeightedPreferences",
            "RelativeSatisfaction",
        ]
        solution_performance = (
            self.student_performance[cols]
            .sum()
            .to_frame()
            .transpose()
            .assign(
                PctAccountedPreferences=lambda df: df["AccountedPreferences"]
                / df["NrPreferences"],
                PctAccountedWeightedPreferences=lambda df: df[
                    "AccountedWeightedPreferences"
                ]
                / df["NrWeightedPreferences"],
            )
        ).to_dict("records")[0]
        return solution_performance

    def _determine_satisfied_preferences_studentindex(self) -> pd.DataFrame:
        """Get the satisfied preferences, but change the index so that it matches the input

        This is useful so that we can match the original file whether a preference is satisfied
        And is used in coloring the output

        Returns
        -------
        df
        """
        preferences_incl_liever_niet = datareader.toggle_negative_weights(
            self.preferences, mask="Gewicht"
        )
        mapping = {}
        for i in range(len(self.preferences)):
            if self.preferences.index[i][1] == "Graag met":
                mapping[self.preferences.reset_index("TypeWens").index[i]] = (
                    preferences_incl_liever_niet.index[i]
                )

        if not mapping:
            return pd.DataFrame(
                columns=["Satisfied", "WeightedSatisfied"],
                index=pd.MultiIndex.from_tuples(
                    [], names=["Leerling", "TypeWens", "Nr"]
                ),
            )

        df = pd.DataFrame(mapping, index=["Leerling", "TypeWens", "Nr"]).transpose()
        df.index.names = ["student", "Nr"]

        df = (
            df.join(self.satisfied_constraints)
            .reset_index(drop=True)
            .set_index(["Leerling", "TypeWens", "Nr"])
        )
        return df

    @staticmethod
    def _display_satisfied_preferences(
        df: pd.DataFrame, satisfied_preferences_original_index: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Determine the background property based on whether a wish is satisfied

        Parameters
        ----------
        df : pd.DataFrame
            DataFrame that contains the right index and columns
        satisfied_preferences_original_index : pd.DataFrame
            DataFrame that contains wether the preference is satisfied. Index in long-form

        """
        df_style = pd.DataFrame(
            "background-color: white",
            index=df.index,
            columns=df.columns,
        )

        for idx in df_style.index:
            for col in df_style.columns:
                original_idx = (idx, col[0], col[1])
                try:
                    if satisfied_preferences_original_index.loc[
                        original_idx, "Satisfied"
                    ]:
                        df_style.loc[idx, col] = "background-color: green"
                    else:
                        df_style.loc[idx, col] = "background-color: red"
                except KeyError:  # Not a preference -> leave background as is
                    continue

        return df_style

    def display_satisfied_preferences(self) -> pd.DataFrame:
        """Display which preferences were satisfied and which werent in the original format

        Returns
        -------
        pd.DataFrame
            Style DataFrame for optimal clarity
        """
        satisfied_preferences_original_index = (
            self._determine_satisfied_preferences_studentindex()
        )

        styled = self.input_sheet.style.apply(
            self._display_satisfied_preferences,
            axis=None,
            satisfied_preferences_original_index=satisfied_preferences_original_index,
        )
        return (
            # Cells (wish targets) and the student index are rendered raw via | safe.
            styled.format(na_rep="", escape="html")
            .format_index(na_rep="", axis="columns")
            .format_index(escape="html", axis="index")
            .set_table_styles(TABLE_STYLES)
        )

    @staticmethod
    def _autoscale_column_width(sheet):
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[get_column_letter(column[0].column)].width = (
                adjusted_width
            )

    def to_excel(self, fname) -> None:
        """Put the most important outcomes of the solution in an Excel file

        Uses the three most important outcomes:
        - The acutal groepsindeling
        - The satisfaction per student
        - Which preferences were accounted for

        Each outcome is styled and shown in its own worksheet
        The solution metrics are not shown - they are probably too abstract for the
        end user

        Parameters
        ----------
        fname : str
            The filename (or file-like object) to write the workbook to
        """
        # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
        with pd.ExcelWriter(fname, engine="openpyxl") as writer:
            self._write_groepsindeling(writer)
            self.group_report.to_excel(writer, sheet_name="Klassenoverzicht")
            self.display_transition_matrix().to_excel(
                writer, sheet_name="Overgangsmatrix"
            )

            self.display_student_performance().to_excel(
                writer, sheet_name="Leerlingtevredenheid"
            )
            sheet = writer.book.worksheets[-1]
            for cell in sheet["B"]:
                cell.number_format = numbers.FORMAT_PERCENTAGE
            self._autoscale_column_width(sheet)

            self.display_satisfied_preferences().to_excel(
                writer, sheet_name="VervuldeWensen"
            )
            sheet = writer.book.worksheets[-1]
            for cell in sheet["B"]:
                cell.number_format = numbers.FORMAT_PERCENTAGE

    def _write_groepsindeling(self, writer):
        groepsindeling = self.display_groepsindeling()
        groepsindeling.iloc[:-1].to_excel(writer, sheet_name="Groepsindeling")
        sheet = writer.sheets["Groepsindeling"]

        row = (
            len(groepsindeling) + len(groepsindeling.columns.levels) + 1
        )  # Excel is 1-based + header
        col_index = 2  # Start bij kolom B in Excel (A is index)
        for group in groepsindeling.columns.levels[0]:
            sheet.merge_cells(
                start_row=row,
                start_column=col_index,
                end_row=row,
                end_column=col_index + 1,
            )
            sheet.cell(row=row, column=col_index).value = groepsindeling.loc[
                "Groepsgrootte", (group, "Jongen")
            ]
            sheet.cell(row=row, column=col_index).alignment = Alignment("center")
            col_index += 2
        self._autoscale_column_width(sheet)

    def get_hash(self) -> int:
        "Give unique code for groepsindeling"
        return hash(tuple(self._get_outcome().sort_values("Naam")["Group"]))

    def compare_to(self, other):
        """Compares this solution to another

        Parameters
        ----------
        other : SolutionAnalyzer object
            Other solution to be compared
        """
        diffs_groepsindeling = (
            self.groepsindeling.set_index("Naam")
            .join(
                other.groepsindeling.set_index("Naam"),
                how="left",
                lsuffix="_this",
                rsuffix="_other",
            )
            .query("Group_this != Group_other")
        )

        for i, row in diffs_groepsindeling.iterrows():
            print(f"{i}:\t{row['Group_this']} --> {row['Group_other']}")
        print("\n" + "-" * 30 + "\n")
        diffs_satisfaction = (
            self.student_performance[["RelativeSatisfaction"]]
            .join(
                other.student_performance[["RelativeSatisfaction"]],
                how="left",
                lsuffix="_this",
                rsuffix="_other",
            )
            .query("RelativeSatisfaction_this != RelativeSatisfaction_other")
        )

        for i, row in diffs_satisfaction.iterrows():
            print(
                f"{i}:\t{row['RelativeSatisfaction_this']:.1%}"
                f" --> {row['RelativeSatisfaction_other']:.1%}"
            )

    def show_all(self, fname="solution.xlsx", to_excel=True):
        """Show all views of the outcome. Only works in Jupyter notebooks"""
        display(self.display_groepsindeling())
        display(self.group_report)
        display(self.display_student_performance())
        display(self.display_transition_matrix())
        display(self.display_satisfied_preferences())
        if to_excel:
            self.to_excel(fname)
